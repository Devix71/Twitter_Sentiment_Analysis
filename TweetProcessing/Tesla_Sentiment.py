import re
import string

import flair
import openpyxl


def deEmojify(text):
    regrex_pattern = re.compile(pattern="["
                                        u"\U0001F600-\U0001F64F"  # emoticons
                                        u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                                        u"\U0001F680-\U0001F6FF"  # transport & map symbols
                                        u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
                                        "]+", flags=re.UNICODE)
    return regrex_pattern.sub(r'', str(text))


def strip_links(text):
    link_regex = re.compile('((https?):((//)|(\\\\))+([\w\d:#@%/;$()~_?\+-=\\\.&](#!)?)*)', re.DOTALL)
    links = re.findall(link_regex, text)
    for link in links:
        text = text.replace(link[0], ', ')
    return text


def strip_all_entities(text):
    entity_prefixes = ['@', '#']
    for separator in string.punctuation:
        if separator not in entity_prefixes:
            text = text.replace(separator, ' ')
    words = []
    for word in text.split():
        word = word.strip()
        if word:
            if word[0] not in entity_prefixes:
                words.append(word)
    return ' '.join(words)


def clean_up(text):
    user.sub('',
             tesla.sub('Tesla',
                       NVDA.sub('NVDIA ',
                                AMD.sub('AMD ', nasdaq.sub
                                ('nasdaq',
                                 web_address.sub('',
                                                 whitespace.sub(' ', text)))))))
    return text


sentiment_model = flair.models.TextClassifier.load('en-sentiment')
whitespace = re.compile(r"\s+")
web_address = re.compile(r"(?i)http(s):\/\/[a-z0-9.~_\-\/]+")
nasdaq = re.compile(r"(?i)@NASDAQ(?=\b)")
NVDA = re.compile(r"(?i)@NVDA(?=\b)")
AMD = re.compile(r"(?i)@AMD(?=\b)")
tesla = re.compile(r"(?i)@Tesla(?=\b)")
user = re.compile(r"(?i)@[a-z0-9_]+")
path = "TSLA.xlsx"
path2 = "TSLA_values.xlsx"
wb = openpyxl.load_workbook(path)
wb2 = openpyxl.load_workbook(path2)
sheet = wb.active
sheet2 = wb2.active
for i in range(1, sheet.max_row + 1):
    sentence = flair.data.Sentence(clean_up(strip_all_entities(strip_links(deEmojify(sheet.cell(i, 6).value)))))
    sentiment_model.predict(sentence)
    sheet2.cell(i, 1).value = sentence.labels[0].score
    sheet2.cell(i, 2).value = sentence.labels[0].value
wb2.save(path2)
print("Parsing complete")
